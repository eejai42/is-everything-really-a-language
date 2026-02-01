#!/usr/bin/env python3
"""
Read test answers from the xlsx substrate.

This script reads the rulebook.xlsx file and populates test-answers.json
by matching xlsx columns to JSON fields.
"""

import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def to_snake_case(name):
    """Convert PascalCase or camelCase to snake_case."""
    s1 = re.sub('(.)([A-Z][a-z]+)', r'\1_\2', name)
    return re.sub('([a-z0-9])([A-Z])', r'\1_\2', s1).lower()


def convert_cell_value(value):
    """Convert Excel cell value to appropriate Python type."""
    if value is None:
        return None
    elif isinstance(value, bool):
        return value
    elif isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value
    elif isinstance(value, str):
        if value.upper() == 'TRUE':
            return True
        elif value.upper() == 'FALSE':
            return False
        elif value.strip() == '':
            return None
        return value
    return value


def fill_null_fields_from_xlsx(xlsx_path, answers_path):
    """Read test-answers.json and fill null fields from xlsx values."""

    with open(answers_path, 'r', encoding='utf-8') as f:
        answers = json.load(f)

    if not answers:
        print("Error: test-answers.json is empty")
        sys.exit(1)

    json_fields = list(answers[0].keys())
    print(f"Found {len(json_fields)} fields in test answers")

    # Find primary key field
    pk_field = next((f for f in json_fields if f.endswith('_id')), json_fields[0])
    print(f"Primary key field: {pk_field}")

    # Load workbook
    wb = load_workbook(xlsx_path)

    # Find worksheet with primary key column
    pk_pascal = ''.join(word.capitalize() for word in pk_field.split('_'))
    ws = None
    for name in wb.sheetnames:
        sheet = wb[name]
        headers = [cell.value for cell in sheet[1] if cell.value]
        if pk_pascal in headers:
            ws = sheet
            print(f"Reading from worksheet: {name}")
            break

    if ws is None:
        print(f"Error: Could not find worksheet with column '{pk_pascal}'")
        sys.exit(1)

    # Get headers and build column mapping
    headers = [cell.value for cell in ws[1] if cell.value]
    column_map = {}
    for header in headers:
        snake = to_snake_case(header)
        if snake in json_fields:
            column_map[header] = snake

    print(f"Found {len(headers)} columns, matched {len(column_map)} to JSON fields")

    # Build lookup from xlsx
    xlsx_lookup = {}
    for row in ws.iter_rows(min_row=2):
        if all(cell.value is None for cell in row[:len(headers)]):
            continue

        xlsx_row = {}
        pk_value = None
        for col_idx, cell in enumerate(row[:len(headers)]):
            header = headers[col_idx]
            if header in column_map:
                json_field = column_map[header]
                value = convert_cell_value(cell.value)
                xlsx_row[json_field] = value
                if json_field == pk_field:
                    pk_value = value

        if pk_value is not None:
            xlsx_lookup[pk_value] = xlsx_row

    print(f"Loaded {len(xlsx_lookup)} rows from xlsx")

    # Fill null fields
    fields_filled = 0
    records_updated = 0

    for answer in answers:
        pk_value = answer.get(pk_field)
        if pk_value is None or pk_value not in xlsx_lookup:
            continue

        xlsx_row = xlsx_lookup[pk_value]
        record_updated = False

        for field in json_fields:
            if answer.get(field) is None and xlsx_row.get(field) is not None:
                answer[field] = xlsx_row[field]
                fields_filled += 1
                record_updated = True

        if record_updated:
            records_updated += 1

    print(f"Filled {fields_filled} null fields across {records_updated} records")

    with open(answers_path, 'w', encoding='utf-8') as f:
        json.dump(answers, f, indent=2)

    print(f"Updated {answers_path}")


def main():
    script_dir = Path(__file__).parent
    xlsx_path = script_dir / 'rulebook.xlsx'
    answers_path = script_dir / 'test-answers.json'

    if not xlsx_path.exists():
        print(f"Error: {xlsx_path} not found")
        sys.exit(1)

    if not answers_path.exists():
        print(f"Error: {answers_path} not found")
        sys.exit(1)

    fill_null_fields_from_xlsx(xlsx_path, answers_path)
    print("xlsx: test-answers.json updated with values from rulebook.xlsx")


if __name__ == "__main__":
    main()
